import streamlit as st
import pandas as pd
import io
import openpyxl
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch

# --- Page Config ---
st.set_page_config(page_title="Excel Filter & Export Tool", layout="wide")

# --- Helper Functions ---

@st.cache_data
def load_data(file, header_row=0, sheet_name=0, handle_merged=False):
    """
    Loads the Excel file into a dataframe.
    If handle_merged is True, it uses openpyxl to unmerge cells and replicate values.
    """
    try:
        if not handle_merged:
            return pd.read_excel(file, header=header_row, sheet_name=sheet_name)
        else:
            # Robust Merged Cell Handling
            wb = openpyxl.load_workbook(file, data_only=True)
            
            # Handle sheet selection
            if isinstance(sheet_name, int):
                ws = wb.worksheets[sheet_name]
            else:
                ws = wb[sheet_name]
            
            # Identify and unmerge cells, replicating the top-left value
            # We copy the ranges to a list first because we are modifying the collection
            for range_ in list(ws.merged_cells.ranges):
                min_col, min_row, max_col, max_row = range_.bounds
                top_left_value = ws.cell(row=min_row, column=min_col).value
                
                # Unmerge
                ws.unmerge_cells(str(range_))
                
                # Fill the unmerged range with the value
                for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                    for cell in row:
                        cell.value = top_left_value
            
            # Extract data manually to create DataFrame
            data = ws.values
            all_rows = []
            
            for i, row in enumerate(data):
                all_rows.append(list(row))
            
            # Slicing for header
            if header_row < len(all_rows):
                # Process headers to handle None values
                raw_headers = all_rows[header_row]
                headers = [str(h) if h is not None else f"Unnamed: {idx}" for idx, h in enumerate(raw_headers)]
                
                # Data rows are everything after header
                df_data = all_rows[header_row+1:]
                return pd.DataFrame(df_data, columns=headers)
            else:
                return pd.DataFrame(all_rows)

    except Exception as e:
        st.error(f"Error loading file: {e}")
        return None

def convert_df_to_excel(df):
    """Converts dataframe to Excel bytes for download."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Filtered_Data')
    return output.getvalue()

def convert_df_to_pdf(df):
    """Converts dataframe to PDF bytes for download using ReportLab."""
    output = io.BytesIO()
    
    # --- Dynamic Page Size Logic ---
    num_cols = len(df.columns)
    min_col_width = 1.5 * inch
    total_page_width = max(11 * inch, num_cols * min_col_width + 1 * inch)
    page_height = 8.5 * inch 
    
    custom_pagesize = (total_page_width, page_height)
    
    doc = SimpleDocTemplate(
        output, 
        pagesize=custom_pagesize,
        leftMargin=0.5*inch,
        rightMargin=0.5*inch,
        topMargin=0.5*inch,
        bottomMargin=0.5*inch
    )
    elements = []
    
    # Add Title
    styles = getSampleStyleSheet()
    title_style = styles['Title']
    title_style.alignment = 1 # Center
    title = Paragraph("Filtered Data Report", title_style)
    elements.append(title)
    
    # Define style for table cells (wraps text)
    cell_style = styles["Normal"]
    cell_style.fontSize = 9
    cell_style.leading = 11
    
    # Prepare Data: Convert everything to Paragraphs
    header = [Paragraph(f"<b>{str(col)}</b>", cell_style) for col in df.columns]
    data = [header]
    
    for _, row in df.iterrows():
        row_data = []
        for item in row:
            text = str(item) if pd.notna(item) else ""
            text = text.replace("\n", "<br/>")
            row_data.append(Paragraph(text, cell_style))
        data.append(row_data)
    
    available_width = total_page_width - 1 * inch 
    col_width = available_width / num_cols
    
    t = Table(data, colWidths=[col_width] * num_cols)
    
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'), 
        ('VALIGN', (0, 0), (-1, -1), 'TOP'), 
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ])
    t.setStyle(style)
    
    elements.append(t)
    
    try:
        doc.build(elements)
    except Exception as e:
        return None 
        
    return output.getvalue()

# --- Main Interface ---

st.title("📊 Excel Data Filter & Export Studio")
st.markdown("Upload your Excel sheet, filter rows/columns, and export the result.")

# 1. File Upload Section
st.header("1. Upload File")
uploaded_file = st.file_uploader("Choose an Excel file", type=["xlsx", "xls"])

if uploaded_file is not None:
    # 1. Get Sheet Names
    try:
        # Load workbook just to get sheet names without loading all data yet
        xl = pd.ExcelFile(uploaded_file)
        sheet_names = xl.sheet_names
    except Exception as e:
        st.error(f"Error reading Excel file structure: {e}")
        st.stop()

    # 2. Settings Columns
    col_settings_1, col_settings_2 = st.columns(2)
    
    with col_settings_1:
        selected_sheet = st.selectbox(
            "Select Worksheet",
            options=sheet_names
        )

    with col_settings_2:
        header_row_index = st.number_input(
            "Select Header Row Index (0 for first row, etc.)",
            min_value=0,
            value=0,
            step=1,
            help="Change this if your column names are not in the first row."
        )

    # 3. Load Data Options
    process_merged_cells = st.checkbox(
        "Process Merged Cells (Unmerge & Fill Down)",
        value=False,
        help="If checked, cells that are merged in Excel will be unmerged, and the value will be copied to all cells in that range. Useful for filtering grouped rows."
    )

    # Reset file pointer to beginning after reading sheet names
    uploaded_file.seek(0)
    
    with st.spinner("Loading data..."):
        df_original = load_data(
            uploaded_file, 
            header_row=header_row_index, 
            sheet_name=selected_sheet, 
            handle_merged=process_merged_cells
        )
    
    if df_original is not None:
        st.success("File uploaded successfully!")
        
        # Initialize the filtered dataframe
        df_filtered = df_original.copy()
        all_columns = df_original.columns.tolist()

        # Layout: Split into sidebar (Filters) and Main Area (Display)
        
        # --- Sidebar: Row Filtering ---
        st.sidebar.header("Filter Rows")
        
        # Reset Button for Row Filters
        if st.sidebar.button("Reset Row Filters"):
            st.session_state["row_filter_cols"] = []
            # Clear specific filter values
            for key in list(st.session_state.keys()):
                if key.startswith("row_val_"):
                    del st.session_state[key]
            st.rerun() # Force a rerun to update UI immediately
        
        st.sidebar.markdown("Select columns to filter by value:")
        
        # Step 1: Choose which columns to apply filters to
        filter_cols = st.sidebar.multiselect(
            "Choose columns to filter rows:",
            options=df_filtered.columns,
            key="row_filter_cols"
        )
        
        # Step 2: Generate dynamic widgets for selected columns
        for col in filter_cols:
            unique_values = df_original[col].dropna().unique() # DropNA to avoid issues with NaN in multiselect
            selected_values = st.sidebar.multiselect(
                f"Select values for '{col}'",
                options=unique_values,
                default=unique_values,
                key=f"row_val_{col}"
            )
            # Apply Filter
            if selected_values:
                df_filtered = df_filtered[df_filtered[col].isin(selected_values)]
            else:
                # If nothing selected, show nothing? Or show all? Usually empty selection means empty result
                df_filtered = df_filtered[df_filtered[col].isin([])] 

        # --- Main Area: Column Selection & Preview ---
        
        col_header, col_reset = st.columns([5, 1])
        with col_header:
            st.header("2. Select Columns")
        with col_reset:
            st.write("") 
            if st.button("Reset Column Selection"):
                st.session_state["final_view_cols"] = all_columns
                st.rerun()

        selected_columns = st.multiselect(
            "Choose which columns to include in the final view:",
            options=all_columns,
            default=all_columns,
            key="final_view_cols"
        )
        
        # Apply Column Selection
        if selected_columns:
            df_final = df_filtered[selected_columns]
        else:
            st.warning("Please select at least one column.")
            df_final = pd.DataFrame()

        # --- Display Data ---
        st.header("3. Data Preview")
        st.write(f"Showing {len(df_final)} rows and {len(df_final.columns)} columns.")
        st.dataframe(df_final, use_container_width=True)

        # --- Export Section ---
        st.header("4. Export Data")
        
        if not df_final.empty:
            file_name_input = st.text_input("Enter file name for export (without extension):", value="filtered_data")
            
            if not file_name_input.strip():
                file_name_input = "filtered_data"
            
            col1, col2 = st.columns(2)
            
            with col1:
                excel_data = convert_df_to_excel(df_final)
                st.download_button(
                    label="📥 Download as Excel (.xlsx)",
                    data=excel_data,
                    file_name=f"{file_name_input}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            with col2:
                if st.button("Generate PDF Preview"):
                    pdf_data = convert_df_to_pdf(df_final)
                    if pdf_data:
                        st.download_button(
                            label="📥 Download as PDF",
                            data=pdf_data,
                            file_name=f"{file_name_input}.pdf",
                            mime="application/pdf"
                        )
                    else:
                        st.error("Could not generate PDF. The table might be too wide or contain incompatible characters.")
        else:
            st.info("No data to export based on current filters.")

else:
    st.info("Awaiting file upload...")